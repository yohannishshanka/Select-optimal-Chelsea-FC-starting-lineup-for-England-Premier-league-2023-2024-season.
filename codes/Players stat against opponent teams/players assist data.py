#!/usr/bin/env python
# coding: utf-8

# In[1]:


import csv
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver

# List of URLs containing the tables
urls = [
   'https://www.transfermarkt.com/fc-arsenal/startseite/verein/11',
    'https://www.transfermarkt.com/manchester-city/startseite/verein/281',
    'https://www.transfermarkt.com/fc-chelsea/startseite/verein/631',
    'https://www.transfermarkt.com/fc-liverpool/startseite/verein/31',
    'https://www.transfermarkt.com/tottenham-hotspur/startseite/verein/148',
    'https://www.transfermarkt.com/manchester-united/startseite/verein/985',
    'https://www.transfermarkt.com/newcastle-united/startseite/verein/762',
    'https://www.transfermarkt.com/aston-villa/startseite/verein/405',
    'https://www.transfermarkt.com/brighton-amp-hove-albion/startseite/verein/1237',
    'https://www.transfermarkt.com/west-ham-united/startseite/verein/379',
    'https://www.transfermarkt.com/fc-brentford/startseite/verein/1148',
    'https://www.transfermarkt.com/crystal-palace/startseite/verein/873',
    'https://www.transfermarkt.com/nottingham-forest/startseite/verein/703',
    'https://www.transfermarkt.com/afc-bournemouth/startseite/verein/989',
    'https://www.transfermarkt.com/fc-everton/startseite/verein/29',
    'https://www.transfermarkt.com/fc-fulham/startseite/verein/931',
    'https://www.transfermarkt.com/wolverhampton-wanderers/startseite/verein/543',
    'https://www.transfermarkt.com/fc-burnley/startseite/verein/1132',
    'https://www.transfermarkt.com/sheffield-united/startseite/verein/350',
    'https://www.transfermarkt.com/luton-town/startseite/verein/1031'
]

chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'

# Replace with the path to your ChromeDriver executable
driver = webdriver.Chrome(executable_path=chrome_driver_path)
# Create an empty dictionary to store player data for each URL
all_player_data = {}

# Create a dictionary to map team names to specific sheet names
team_sheet_mapping = {
    
    'fc-arsenal': 'Arsenal',
    'manchester-city': 'Manchester City',
    'fc-chelsea': 'Chelsea',
    'fc-liverpool': 'Liverpool',
    'tottenham-hotspur': 'Tottenham Hotspur',
    'manchester-united': 'Manchester United',
    'newcastle-united': 'Newcastle United',
    'aston-villa': 'Aston Villa',
    'brighton-amp-hove-albion': 'Brighton & Hove Albion',
    'west-ham-united': 'West Ham United',
    'fc-brentford': 'FC Brentford',
    'crystal-palace': 'Crystal Palace',
    'nottingham-forest': 'Nottingham Forest',
    'afc-bournemouth': 'AFC Bournemouth',
    'fc-everton': 'FC Everton',
    'fc-fulham': 'FC Fulham',
    'wolverhampton-wanderers': 'Wolverhampton Wanderers',
    'fc-burnley': 'FC Burnley',
    'sheffield-united': 'Sheffield United',
    'luton-town': 'Luton Town'
  
}

for url in urls:
    # Load the page using Selenium
    driver.get(url)

    # Now, fetch the page content after a brief wait
    driver.implicitly_wait(5)
    page_content = driver.page_source

    # Parse the HTML content of the page
    soup = BeautifulSoup(page_content, 'html.parser')

    # Specify the table name
    table_name = 'items'

    # Find the table element based on its class name
    table = soup.find('table', class_='items')

    # Check if the table is found
    if table:
        # Create a list to store the data
        player_data = []

        # Extract and append player names based on positions
        positions = [  "Right Winger", "Left Winger", "Defensive Midfield", "Left-Back", "Attacking Midfield", "Centre-Back", "Right-Back", "Central Midfield", "Centre-Forward", "Second Striker", "Left Midfield", "Right Midfield"]


        for position in positions:
            for row in table.find_all('tr')[1:]:  # Skip the header row
                if position in row.get_text():
                    player_name_element = row.find('td', class_='hauptlink')
                    player_name = player_name_element.text.strip() if player_name_element else None
                    if player_name is not None:
                        player_data.append({"Player": player_name, "Position": position})

        # Store player data in the dictionary with URL as key
        all_player_data[url] = player_data
    else:
        print(f"No table with class '{table_name}' found on the page for URL: {url}")

# Close the browser window after processing all URLs
driver.quit()

# Save data to the same Excel file with different sheets
excel_file_path = 'players assist data.xlsx'  # Use an Excel file for multiple sheets
with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
    for url, player_data in all_player_data.items():
        # Convert player data to a DataFrame
        df = pd.DataFrame(player_data, columns=["Player", "Position"])

        # Extract the team name from the URL
        team_name = url.split("/")[3]

        # Get the corresponding sheet name from the mapping dictionary
        sheet_name = team_sheet_mapping.get(team_name, team_name)

        # Write the DataFrame to a sheet with the specific name
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"Data saved to {excel_file_path}")

import openpyxl
from unidecode import unidecode

# Open the existing Excel file
file_path = 'players assist data.xlsx'
workbook = openpyxl.load_workbook(file_path)

# Iterate over each sheet in the workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    
    # Find the index of the 'Player' column
    player_column_index = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'Player':
            player_column_index = col
            break
    
    # If 'Player' column is found, apply Unicode transformation and save it to a new column
    if player_column_index:
        # Calculate the index for the new column (3rd column)
        new_column_index = 3
        
        # Iterate over each row, starting from the second row (as the first row contains headers)
        for row in range(2, sheet.max_row + 1):
            # Get the value of the 'Player' column in the current row
            player_value = sheet.cell(row=row, column=player_column_index).value
            
            # Apply Unicode transformation using unidecode
            unicode_player_value = unidecode(str(player_value))
            
            unicode_player_value =unicode_player_value.replace("-"," ")
            
            # Write the Unicode-transformed value to the new column
            sheet.cell(row=row, column=new_column_index, value=unicode_player_value)
    
# Save the modified workbook to a new Excel file
new_file_path = 'players assist data.xlsx'
workbook.save(new_file_path)

# Close the workbook
workbook.close()

from openpyxl import load_workbook

# Load the workbook
workbook = load_workbook('players assist data.xlsx')

worksheet_names = workbook.sheetnames

# Loop through each worksheet in the workbook
for sheet_name in worksheet_names:
    ws = workbook[sheet_name]
    for i, col_name in enumerate(worksheet_names, start=4):
        if i > 23:  # Stop renaming after the 22nd column
            break
        col_letter = ws.cell(row=1, column=i).column_letter
        ws[col_letter + '1'] = col_name

# Save the modified workbook
workbook.save('players assist data.xlsx')

workbook = load_workbook('players assist data.xlsx')

# Get the sheet names
worksheet_names = workbook.sheetnames

# Iterate over each sheet and delete it if the name is not 'Chelsea'
for sheet_name in worksheet_names:
    if sheet_name != 'Chelsea':
        workbook.remove(workbook[sheet_name])

# Save the modified workbook
workbook.save('players assist data.xlsx')

import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from unidecode import unidecode 
import openpyxl

# URLs to scrape
urls = [
   
    'https://www.transfermarkt.com/fc-chelsea/startseite/verein/631',
    
]

chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'

# Replace with the path to your ChromeDriver executable
driver = webdriver.Chrome(executable_path=chrome_driver_path)


# Initialize an empty dictionary to store filtered table data URLs wise
filtered_table_data_dict = {}

# Desired positions
desired_positions = [
    "Goalkeeper", "Right Winger", "Left Winger", "Defensive Midfield", 
    "Left-Back", "Attacking Midfield", "Centre-Back", "Right-Back", 
    "Central Midfield", "Centre-Forward", "Second Striker", 
    "Left Midfield", "Right Midfield"
]

# Iterate through each URL
for url in urls:
    # Open the URL
    driver.get(url)

    # Get the HTML content of the page
    html = driver.page_source

    # Use BeautifulSoup to parse the HTML
    soup = BeautifulSoup(html, 'html.parser')

    # Find the table with class name "items"
    items_table = soup.find('table', class_='items')
    
    # Find all table rows
    rows = items_table.find_all('tr')

    # Initialize an empty list to store filtered table data for this URL
    filtered_table_data = []

    # Iterate through each row
    for row in rows:
        # Find all cells in the row
        cells = row.find_all('td')
        
        # Check if there are cells in the row
        if cells:
            # Extract text from the first cell (position) of the row
            position_text = cells[0].get_text(strip=True)
            
            # Check if the position is in the desired positions list
            if position_text in desired_positions:
                # Extract text from each cell and append to the filtered list
                row_data = [cell.get_text(strip=True) for cell in cells]
                filtered_table_data.append(row_data)
    
    # Store the filtered data for this URL in the dictionary
    filtered_table_data_dict[url] = filtered_table_data
    
    for url, data in filtered_table_data_dict.items():
        defender_indices = [index for index, row in enumerate(data) if row[0] ==  "Left-Back"]
        
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in defender_indices:
            a.append(defender_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    defender_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in defender_indices:
        defender_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in defender_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        defender_indices = [index for index, row in enumerate(data) if row[0] ==  "Centre-Back"]
        #print("Goalkeeper Indices:", goalkeeper_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in defender_indices:
            a.append(defender_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    defender_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in defender_indices:
        defender_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in defender_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in  (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        defender_indices = [index for index, row in enumerate(data) if row[0] == "Right-Back"]
        #print("Goalkeeper Indices:", goalkeeper_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in defender_indices:
            a.append(defender_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    defender_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in defender_indices:
        defender_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in defender_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in  (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        midfielder_indices = [index for index, row in enumerate(data) if row[0] ==  "Defensive Midfield"]
        #print("Goalkeeper Indices:", goalkeeper_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in midfielder_indices:
            a.append(midfielder_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    midfielder_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in midfielder_indices:
        midfielder_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in midfielder_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        midfielder_indices = [index for index, row in enumerate(data) if row[0] == "Central Midfield"]
        #print("Goalkeeper Indices:", goalkeeper_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in midfielder_indices:
            a.append(midfielder_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    midfielder_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in midfielder_indices:
        midfielder_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in midfielder_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in  (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        midfielder_indices = [index for index, row in enumerate(data) if row[0] ==  "Attacking Midfield"]
        #print("Goalkeeper Indices:", goalkeeper_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in midfielder_indices:
            a.append(midfielder_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    midfielder_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in midfielder_indices:
        midfielder_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in midfielder_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in  (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        midfielder_indices = [index for index, row in enumerate(data) if row[0] ==  "Left Midfield"]
        #print("Goalkeeper Indices:", midfielder_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in midfielder_indices:
            a.append(midfielder_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    midfielder_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in midfielder_indices:
        midfielder_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in midfielder_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        midfielder_indices = [index for index, row in enumerate(data) if row[0] ==  "Right Midfield"]
        #print("Goalkeeper Indices:", goalkeeper_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in midfielder_indices:
            a.append(midfielder_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    midfielder_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in midfielder_indices:
        midfielder_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in midfielder_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        defender_indices = [index for index, row in enumerate(data) if row[0] ==  "Left-Back"]
        
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in defender_indices:
            a.append(defender_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    defender_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in defender_indices:
        defender_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in defender_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        defender_indices = [index for index, row in enumerate(data) if row[0] ==  "Centre-Back"]
        #print("Goalkeeper Indices:", goalkeeper_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in defender_indices:
            a.append(defender_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    defender_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in defender_indices:
        defender_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in defender_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in  (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        defender_indices = [index for index, row in enumerate(data) if row[0] == "Right-Back"]
        #print("Goalkeeper Indices:", goalkeeper_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in defender_indices:
            a.append(defender_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    defender_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in defender_indices:
        defender_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in defender_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in  (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        forward_indices = [index for index, row in enumerate(data) if row[0] ==  "Centre-Forward"]
        #print("Goalkeeper Indices:", forward_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in forward_indices:
            a.append(forward_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    forward_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in forward_indices:
        forward_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in forward_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        forward_indices = [index for index, row in enumerate(data) if row[0] ==  "Second Striker"]
        #print("Goalkeeper Indices:", forward_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in forward_indices:
            a.append(forward_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    forward_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in forward_indices:
        forward_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in forward_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in  (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        forward_indices = [index for index, row in enumerate(data) if row[0] ==  "Right Winger"]
        #print("Goalkeeper Indices:", forward_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in forward_indices:
            a.append(forward_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    forward_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in forward_indices:
        forward_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in forward_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in  (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')
            
    for url, data in filtered_table_data_dict.items():
        forward_indices = [index for index, row in enumerate(data) if row[0] ==  "Left Winger"]
        #print("Goalkeeper Indices:", forward_indices)
        
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        
        for index in forwardr_indices:
            a.append(forward_indices[index])
        

    # Pri# Create a new list to store href links for goalkeepers
      # Create a new list to store href links for goalkeepers
    forward_href_list = []

      # Iterate through each index in goalkeeper_indices
    for index in forward_indices:
        forward_href_list.append(filtered_href_list[index])
        
    # Append the corresponding href link from filtered_href_list to goalkeeper_href_list
    #print(goalkeeper_href_list)
          
        
    # Iterate through each filtered href link
    first_parts_modified = []
    new_urls = []
    first_parts_modified_new = []
    

    # Iterate through each filtered href link
    for href in forward_href_list:
        # Split the href link by "/"
        parts = href.split("/")

        # Extract relevant parts
        first_part = parts[1]
        fourth_part = parts[4]

        # Replace "-" with " " in the first part
        first_part_modified = first_part.replace("-", " ")
    
        # Create the new URL
        new_url = f"https://www.transfermarkt.com/{first_part}/bilanz/spieler/{fourth_part}/plus/0?wettbewerb=GB1"

        # Append modified parts to the lists
        first_parts_modified.append(first_part_modified)
        new_urls.append(new_url)
        
    for element in first_parts_modified:
        first_parts_modified_new.append(' '.join(word.capitalize() for word in element.split()))
    

 
    # Print modified parts
    #print(first_parts_modified_new)
    #print(new_urls)
    
    # Iterate through each new URL and scrape data
    for new_url, player_name in zip(new_urls, first_parts_modified_new):
        driver.get(new_url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        table = soup.find('table', class_='items')

        # Extract table rows
        if table is None:
            print(f"No table found with class 'items' in the URL: {new_url}")
            continue

       # Debugging statement to check if the table is found
       # print(table)

       # Extract table rows
        rows = table.find_all('tr')

        if len(rows) <= 1:
            print(f"No data rows found in the table for URL: {new_url}")
            continue
        
        # Initialize an empty dictionary to store the results
        data_dict = {}

        # Loop through rows and print data
        for row in rows[1:]:
            # Skip the first row
                   # Extract the second and seventh elements from each row
            cells = row.find_all(['th', 'td'])
            row_data = [cells[1].get_text(strip=True), cells[7].get_text(strip=True)]
            row_data = ['Manchester United' if cell == 'MUN' else cell for cell in row_data]
            row_data = ['FC Fulham' if cell == 'FUL' else cell for cell in row_data]
            row_data = ['FC Everton' if cell == 'EVE' else cell for cell in row_data]
            row_data = ['Wolverhampton Wanderers' if cell == 'WOL' else cell for cell in row_data]
            row_data = ['Arsenal' if cell == 'ARS' else cell for cell in row_data]
            row_data = ['Manchester City' if cell == 'MCI' else cell for cell in row_data]
            row_data = ['Chelsea' if cell == 'CHE' else cell for cell in row_data]
            row_data = ['Liverpool' if cell == 'LIV' else cell for cell in row_data]
            row_data = ['Tottenham Hotspur' if cell == 'TOT' else cell for cell in row_data]
            row_data = ['Newcastle United' if cell == 'NEW' else cell for cell in row_data]
            row_data = ['Aston Villa' if cell == 'AVL' else cell for cell in row_data]
            row_data = ['Brighton & Hove Albion' if cell == 'BHA' else cell for cell in row_data]
            row_data = ['West Ham United' if cell == 'WHU' else cell for cell in row_data]
            row_data = ['Crystal Palace' if cell == 'CRY' else cell for cell in row_data]
            row_data = ['Nottingham Forest' if cell == 'FOR' else cell for cell in row_data]
            row_data = ['AFC Bournemouth' if cell == 'BOU' else cell for cell in row_data]
            row_data = ['FC Burnley' if cell == 'BUR' else cell for cell in row_data]
            row_data = ['Sheffield United' if cell == 'SHU' else cell for cell in row_data]
            row_data = ['FC Brentford' if cell == 'BRE' else cell for cell in row_data]
            row_data = ['Luton Town' if cell == 'LUT' else cell for cell in row_data]
            # Store the data in the dictionary
            data_dict[row_data[0]] = row_data[1]
            
            import openpyxl

            # Load the existing workbook
            workbook = openpyxl.load_workbook('players assist data.xlsx')

           # Iterate over each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Iterate over each row in the worksheet
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):  # Start index from 1
                          # Check if 'Erling Haaland' is in the first column (assuming first column has player names)
                    if player_name in (row[0], row[2]):
                    # Update columns based on data_dict
                        for header, value in data_dict.items():
                              # Find the column index corresponding to the header
                            for col_index, col_header in enumerate(ws[1], start=1):  # Assuming headers are in the first row
                                if col_header.value == header:
                                # Enter the value into the corresponding column
                                    ws.cell(row=row_index, column=col_index, value=value)
                                    break  # Stop searching for the column index once found

             # Save the workbook
            workbook.save('players assist data.xlsx')


            


        
# Close the webdriver
driver.quit()

workbook = load_workbook('players assist data.xlsx')

# Get the first sheet
sheet = workbook.active

# Remove the third column (column 'C')
sheet.delete_cols(3)

# Save the workbook
workbook.save('players assist data.xlsx')


# In[2]:





# In[ ]:




