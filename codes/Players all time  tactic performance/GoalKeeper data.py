#!/usr/bin/env python
# coding: utf-8

# In[1]:




import csv
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver

# List of URLs containing the tables
urls = [

    'https://www.transfermarkt.com/fc-chelsea/startseite/verein/631',

]

chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'

# Replace with the path to your ChromeDriver executable
driver = webdriver.Chrome(executable_path=chrome_driver_path)


# Create an empty dictionary to store player data for each URL
all_player_data = {}

# Create a dictionary to map team names to specific sheet names
team_sheet_mapping = {
   
    'fc-chelsea': 'Chelsea'}


for url in urls:
    # Load the page using Selenium
    driver.get(url)

    # Now, fetch the page content after a brief wait
    driver.implicitly_wait(5)
    page_content = driver.page_source

    # Parse the HTML content of the page
    soup = BeautifulSoup(page_content, 'html.parser')

    # Specify the table name
    table_name = 'items'

    # Find the table element based on its class name
    table = soup.find('table', class_='items')

    # Check if the table is found
    if table:
        # Create a list to store the data
        player_data = []

        # Extract and append player names based on positions
        positions = ["Goalkeeper"]

        for position in positions:
            for row in table.find_all('tr')[1:]:  # Skip the header row
                if position in row.get_text():
                    player_name_element = row.find('td', class_='hauptlink')
                    player_name = player_name_element.text.strip() if player_name_element else None
                    if player_name is not None:
                        player_data.append({"Player": player_name, "Position": position})

        # Store player data in the dictionary with URL as key
        all_player_data[url] = player_data
    else:
        print(f"No table with class '{table_name}' found on the page for URL: {url}")

# Close the browser window after processing all URLs
driver.quit()

# Save data to the same Excel file with different sheets
excel_file_path = 'Goalkeeper_data.xlsx'  # Use an Excel file for multiple sheets
with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
    for url, player_data in all_player_data.items():
        # Convert player data to a DataFrame
        df = pd.DataFrame(player_data, columns=["Player", "Position"])

        # Extract the team name from the URL
        team_name = url.split("/")[3]

        # Get the corresponding sheet name from the mapping dictionary
        sheet_name = team_sheet_mapping.get(team_name, team_name)

        # Write the DataFrame to a sheet with the specific name
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"Data saved to {excel_file_path}")

import pandas as pd
import os
import unicodedata

# Path to the Excel file
excel_file_path = 'Goalkeeper_data.xlsx'
modified_excel_file_path = 'modified_player_data_performance.xlsx'

# Read the Excel file
xls = pd.ExcelFile(excel_file_path)

# Read data from each sheet, modify it, and save it back
with pd.ExcelWriter(modified_excel_file_path, engine='xlsxwriter') as writer:
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name)
        
        # Create the new column "clean Player" by replacing specific characters and removing diacritics
        df['clean Player'] = df['Player'].apply(lambda x: ''.join(
            unicodedata.normalize('NFKD', c).encode('ASCII', 'ignore').decode('utf-8')
            if c != '-' else ' '
            for c in x
        ))
        replacements = {'Marek Rodak': 'Marek Rodák',"Hakon Rafn Valdimarsson":"Hákon Valdimarsson","Alisson":"Alisson Becker"}
  # Specify the replacements you want to make
        df['clean Player'] = df['clean Player'].replace(replacements, regex=True)
        
        # Write the modified DataFrame to the same sheet
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# Close the original Excel file
xls.close()

# Replace the original file with the modified one
os.replace(modified_excel_file_path, excel_file_path)



from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'

# Replace with the path to your ChromeDriver executable
driver = webdriver.Chrome(executable_path=chrome_driver_path)


# Extract href attributes from club boxes and add "https://www.premierleague.com/" to the beginning
club_urls = ["https://www.premierleague.com/clubs/4/Chelsea/overview"]

def scrape_player_stats(url, driver):
    driver.get(url)
    
    # Click the "Accept All Cookies" button if it's present
    try:
        accept_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
        )
        accept_btn.click()
    except Exception as e:
        print("Accept button not found or clicked:", e)
        
    time.sleep(5)    
        
    try:
        close_button_xpath = '//*[@id="advertClose"]'
        close_button = driver.find_element(By.XPATH ,close_button_xpath)
        close_button.click()

        
    except Exception as e:
        print("Accept button not found or clicked:", e)

    # Find squad link and click it
    squad_link = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, 'li.tab[data-nav-index="1"] a.club-navigation__link[data-link-index="1"]'))
    )
    squad_link.click()

    # Get page source after dynamic content is loaded
    page_source = driver.page_source

    # Parse the HTML using BeautifulSoup
    soup = BeautifulSoup(page_source, "html.parser")

    href_links = []

    # Find all anchor tags (links) and extract href attributes
    links = soup.find_all("a")
    for link in links:
        href = link.get("href")
        if href and href.startswith("/players/"):  # Check if href starts with "/players/"
            href_links.append(href)

    result_dict = {}

    # Assuming href_links is your list of URLs
    for url in href_links:
        # Split the URL by "/"
        parts = url.split("/")

        # Extract relevant parts
        second_part = parts[2]  # Get the second part after splitting
        third_part = parts[3]   # Get the third part after splitting

        # Create new variables based on extracted parts
        new_variable1 = "https://www.premierleague.com/players/" + second_part + "/" + third_part + "/stats"

        # Replace "-" with " " in the second part
        third_part_replaced = third_part.replace("-", " ")

        # Add entry to the dictionary
        result_dict[third_part_replaced] = new_variable1

    # Load the Excel workbook
    workbook = load_workbook("Goalkeeper_data.xlsx")

    # Iterate over each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Rename the first cell in the 13th column to "Yohan"
        sheet.cell(row=1, column=13, value="Total saves per game")
        sheet.cell(row=1, column=4, value="Penalty saved")
        sheet.cell(row=1, column=5, value="Total punches per game")
        sheet.cell(row=1, column=6, value="High claims per game")
        sheet.cell(row=1, column=7, value="Catches per game")
        sheet.cell(row=1, column=8, value="Sweeper clearence per game")
        sheet.cell(row=1, column=9, value="Accurate long balls per game")
        sheet.cell(row=1, column=10, value="Error leading to a goals")
        sheet.cell(row=1, column=11, value="Own goals")
        sheet.cell(row=1, column=12, value="Passes per game")
        
        # Iterate over each row in the sheet
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for key in result_dict.keys():
                if key in row:  # Check if the key is present in the current row
                    url = result_dict[key]
                    driver.get(url)
                    page_source = driver.page_source
                    soup = BeautifulSoup(page_source, "html.parser")

                    apperence_container = soup.find("span", class_="allStatContainer js-all-stat-container statappearances")
                    if apperence_container:
                        apperence = int(apperence_container.text.strip())
                    else:
                        apperence = 0

                        

                    stats_container = soup.find("span", class_="allStatContainer js-all-stat-container statsaves")
                    if stats_container:
                        stats_text = int(stats_container.text.strip().replace(",", ""))
                        if apperence != 0:
                            total_tackle = "{:.2f}".format(stats_text / apperence)
                            sheet.cell(row=row_num, column=13, value=total_tackle)
                        else:
                            print(f"Player: {key}, Apperence is 0")
                            sheet.cell(row=row_num, column=13, value="-")

                    else:
                        print(f"Player: {key}, total saves per game stats not found")
                        

                    stats_container = soup.find("span", class_="allStatContainer js-all-stat-container statpunches")
                    if stats_container:
                        stats_text = int(stats_container.text.strip())
                        if apperence != 0:
                            statpunches = "{:.2f}".format(stats_text / apperence)
                            sheet.cell(row=row_num, column=5, value=statpunches)
                        else:
                            print(f"Player: {key}, Apperence is 0")
                            sheet.cell(row=row_num, column=5, value="-")

                    else:
                        print(f"Player: {key}, punches per game stats not found")
                        

                    stats_container = soup.find("span", class_="allStatContainer js-all-stat-container statpenalty_save")
                    if stats_container:
                        stats_text = stats_container.text.strip()

                        sheet.cell(row=row_num, column=4, value=stats_text)

                    else:
                        print(f"Player: {key}, stat penalty save stats not found")
                        
                        

                    stats_container = soup.find("span", class_="allStatContainer js-all-stat-container statgood_high_claim")
     
                    if stats_container:
                        stats_text = int(stats_container.text.strip().replace(",", ""))
                        if apperence != 0:
                            high_claim = "{:.2f}".format(stats_text / apperence)
                            sheet.cell(row=row_num, column=6, value=high_claim)
                        else:
                            print(f"Player: {key}, Apperence is 0")
                            sheet.cell(row=row_num, column=6, value="-")

                    else:
                        print(f"Player: {key},high claim per game stats not found")
                        

                    stats_container = soup.find("span", class_="allStatContainer js-all-stat-container statcatches")
                    if stats_container:
                        stats_text = int(stats_container.text.strip())
                        if apperence != 0:
                            catches = "{:.2f}".format(stats_text / apperence)
                            sheet.cell(row=row_num, column=7, value=catches)
                        else:
                            print(f"Player: {key}, Apperence is 0")
                            sheet.cell(row=row_num, column=7, value="-")

                    else:
                        print(f"Player: {key}, catches per game stats not found")

                    stats_container = soup.find("span", class_="allStatContainer js-all-stat-container stattotal_keeper_sweeper")

                    if stats_container:
                        stats_text = int(stats_container.text.strip().replace(",", ""))
                        if apperence != 0:
                            sweeper = "{:.2f}".format(stats_text / apperence)
                            sheet.cell(row=row_num, column=8, value=sweeper)
                        else:
                            print(f"Player: {key}, Apperence is 0")
                            sheet.cell(row=row_num, column=8, value="-")

                    else:
                        print(f"Player: {key}, total sweeper clearance per game stats not found")
                        

                    stats_container = soup.find("span", class_="allStatContainer js-all-stat-container stataccurate_long_balls")
                   
                    if stats_container:
                        stats_text = int(stats_container.text.strip().replace(",", ""))
                        if apperence != 0:
                            accurate_long_balls = "{:.2f}".format(stats_text / apperence)
                            sheet.cell(row=row_num, column=9, value=accurate_long_balls)
                        else:
                            print(f"Player: {key}, Apperence is 0")
                            sheet.cell(row=row_num, column=9, value="-")

                    else:
                        print(f"Player: {key}, accurate long balls per game stats not found")
                        

                    stats_container = soup.find("span", class_="allStatContainer js-all-stat-container staterror_lead_to_goal")
                    if stats_container:
                        stats_text = stats_container.text.strip()

                        sheet.cell(row=row_num, column=10, value=stats_text)

                    else:
                        print(f"Player: {key}, error_lead_to_goal stats not found")
                        

                        
                    stats_container = soup.find("span", class_="allStatContainer js-all-stat-container statown_goals")
                    if stats_container:
                        pass_per_game = stats_container.text.strip()

                        sheet.cell(row=row_num, column=11, value=pass_per_game)

                    else:
                        print(f"Player: {key}, Own goals stats not found")
                        
                    stats_container = soup.find("span", class_="allStatContainer js-all-stat-container stattotal_pass_per_game")
                    if stats_container:
                        pass_per_game = stats_container.text.strip()

                        sheet.cell(row=row_num, column=12, value=pass_per_game)

                    else:
                        print(f"Player: {key}, Passes per game stats not found")   
                        
                          
                    

                    break  # Move to the next sheet once a match is found

    # Save the workbook
    workbook.save("Goalkeeper_data.xlsx")



chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'

# Replace with the path to your ChromeDriver executable
driver = webdriver.Chrome(executable_path=chrome_driver_path)

# Iterate over each URL and perform scraping
for url in club_urls:
    scrape_player_stats(url, driver)

# Quit the webdriver
driver.quit()

from openpyxl import load_workbook

# Load the Excel workbook
workbook = load_workbook("Goalkeeper_data.xlsx")

# Iterate over each sheet in the workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == "clean Player":
                column_to_delete = cell.column
                sheet.delete_cols(column_to_delete)

# Save the workbook
workbook.save("Goalkeeper_data.xlsx")


# In[ ]:




