from selenium import webdriver
from bs4 import BeautifulSoup
import pandas as pd

# List of URLs without "platzierungen/"
urls = [
    'https://www.transfermarkt.com/fc-arsenal/startseite/verein/11',
    'https://www.transfermarkt.com/manchester-city/startseite/verein/281',
    'https://www.transfermarkt.com/fc-chelsea/startseite/verein/631',
    'https://www.transfermarkt.com/fc-liverpool/startseite/verein/31',
    'https://www.transfermarkt.com/tottenham-hotspur/startseite/verein/148',
    'https://www.transfermarkt.com/manchester-united/startseite/verein/985',
    'https://www.transfermarkt.com/newcastle-united/startseite/verein/762',
    'https://www.transfermarkt.com/aston-villa/startseite/verein/405',
    'https://www.transfermarkt.com/brighton-amp-hove-albion/startseite/verein/1237',
    'https://www.transfermarkt.com/west-ham-united/startseite/verein/379',
    'https://www.transfermarkt.com/fc-brentford/startseite/verein/1148',
    'https://www.transfermarkt.com/crystal-palace/startseite/verein/873',
    'https://www.transfermarkt.com/nottingham-forest/startseite/verein/703',
    'https://www.transfermarkt.com/afc-bournemouth/startseite/verein/989',
    'https://www.transfermarkt.com/fc-everton/startseite/verein/29',
    'https://www.transfermarkt.com/fc-fulham/startseite/verein/931',
    'https://www.transfermarkt.com/wolverhampton-wanderers/startseite/verein/543',
    'https://www.transfermarkt.com/fc-burnley/startseite/verein/1132',
    'https://www.transfermarkt.com/sheffield-united/startseite/verein/350',
    'https://www.transfermarkt.com/luton-town/startseite/verein/1031']

# Append "platzierungen/" to each URL
new_urls = [url.replace('/startseite/', '/platzierungen/') for url in urls]

# Function to scrape table data
def scrape_table_data(url):
    chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'

    driver = webdriver.Chrome(executable_path=chrome_driver_path)
    driver.get(url)
    driver.implicitly_wait(20)

    soup = BeautifulSoup(driver.page_source, 'html.parser')
    table = soup.find('table', class_='items')

    # Extract the text from the 1st, 3rd, and 7th columns for rows 1 to 15
    data = []
    for index, row in enumerate(table.find_all('tr')):
        if index > 0 and index <= 11:  # Skip the first row (column headers) and get rows 1 to 15
            cells = row.find_all(['th', 'td'])
            first_column = cells[0].get_text(strip=True)
            third_column = cells[2].get_text(strip=True)
            seventh_column = cells[10].get_text(strip=True)
            data.append({'season': first_column, 'Tier': third_column, 'league position': seventh_column})

    # Create a DataFrame from the list of dictionaries
    df = pd.DataFrame(data)

    driver.quit()

    return df

# Create a Pandas Excel writer using XlsxWriter as the engine
writer = pd.ExcelWriter('football_data.xlsx', engine='xlsxwriter')

# Scrape data for each URL and save to a separate sheet
for i, url in enumerate(new_urls, start=1):
    print(f"Scraping data from: {url}")
    df = scrape_table_data(url)
    df.to_excel(writer, sheet_name=f'Sheet{i}', index=False)

# Save the Excel file
writer.save()

from openpyxl import load_workbook

workbook = load_workbook('football_data.xlsx')

# Rename the sheets
sheet_names = ['Arsenal', 'Manchester City', 'Chelsea', 'Liverpool', 'Tottenham Hotspur', 'Manchester United', 'Newcastle United',
              'Aston Villa','Brighton & Hove Albion','West Ham United','FC Brentford','Crystal Palace','Nottingham Forest','AFC Bournemouth',
              'FC Everton','FC Fulham','Wolverhampton Wanderers','FC Burnley','Sheffield United','Luton Town']
for i, sheet_name in enumerate(sheet_names, start=1):
    workbook[f'Sheet{i}'].title = sheet_name
    
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    # Delete rows starting from the second row if the sheet is not empty
    if len(sheet['A']) > 2:
        sheet.delete_rows(2)

# Save the workbook
workbook.save('football_data.xlsx')


