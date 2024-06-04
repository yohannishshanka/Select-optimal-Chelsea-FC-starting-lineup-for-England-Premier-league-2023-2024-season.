#!/usr/bin/env python
# coding: utf-8

# In[1]:


from selenium import webdriver
from bs4 import BeautifulSoup
import openpyxl
import re
from openpyxl import load_workbook

url = "https://www.whoscored.com/Regions/252/Tournaments/2/Seasons/9618/Stages/22076/TeamStatistics/England-Premier-League-2023-2024"

chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'


# Replace with the path to your ChromeDriver executable
driver = webdriver.Chrome(executable_path=chrome_driver_path)
driver.get(url)

# Wait for the page to load completely
driver.implicitly_wait(10)

# Get the page source and create a BeautifulSoup object
soup = BeautifulSoup(driver.page_source, "html.parser")

# Find all <a> tags with href containing "/Teams/"
links = set()
for a in soup.find_all("a", href=True):
    if "/Teams/" in a["href"]:
        links.add("https://www.whoscored.com/" + a["href"])

# Close the Selenium WebDriver
driver.quit()

chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'

# Replace with the path to your ChromeDriver executable
driver = webdriver.Chrome(executable_path=chrome_driver_path)

# Create a new Excel workbook
workbook = openpyxl.Workbook()

for url in links:
    # Load the webpage
    driver.get(url)

    driver.implicitly_wait(10)  # Adjust the timeout as needed

    # Get the page source
    page_source = driver.page_source

    # Parse the page source with BeautifulSoup
    soup = BeautifulSoup(page_source, "html.parser")

    # Find all elements with the specified class
    elements = soup.find_all("div", class_="sws-content character-card singular")

    # Extract and store text before "Style of Play" from each element
    row = 2  # Start from row 2 (after the headers)
    sheet_name = url.split("England-")[1]  # Extract sheet name from URL
    sheet = workbook.create_sheet(title=sheet_name)  # Create a new sheet

    # Add column headers
    sheet["A1"] = "KPIs"
    sheet["B1"] = "Strength/Weakness"
    
    start_index = url.find("England-") + len("England-")

    # Extract the part of the URL after "England-"
    after_england = url[start_index:]

    # Count the occurrences of "-" in the extracted part
    count = after_england.count("-")
    
    if count==0:
        for element in elements:
            text = element.get_text()
            # Remove words after "Style of Play"
            style_of_play_index = text.find("Style of Play")
            if style_of_play_index != -1:
                text = text[:style_of_play_index]
                # Remove two words before "Style of Play"
                words = text.split()
            if len(words) > 2:
                text = " ".join(words[:-1])
            # Remove "Strengths", "Weaknesses", "+", "-"
            text = text.replace("Strengths", "").replace("Weaknesses", "").replace("+", "").replace("-", "")
            # Add "-" before and after "Very Strong", "Strong", or "Weak"
            text = text.replace("Very Strong", "-Very Strong-").replace("Very Weak","-Very Weak-").replace("Strong", "-Strong-").replace("Weak", "-Weak-").replace("Very Weak","-Very Weak-")
            text = text.replace("Very -Strong-", "Very Strong").replace("Very -Weak-","Very Weak")
            # Split the text by "-" and write to the Excel file
            parts = text.split("-")
            for i in range(0, len(parts), 2):
                sheet.cell(row, 1, parts[i].strip())
                if i + 1 < len(parts):
                    sheet.cell(row, 2, parts[i + 1].strip())
                row += 1
                
    else:
        for element in elements:
            text = element.get_text()
          # Remove words after "Style of Play"
            style_of_play_index = text.find("Style of Play")
            if style_of_play_index != -1:
                text = text[:style_of_play_index]
            # Remove two words before "Style of Play"
                words = text.split()
            if len(words) > 2:
                text = " ".join(words[:-2])
            # Remove "Strengths", "Weaknesses", "+", "-"
            text = text.replace("Strengths", "").replace("Weaknesses", "").replace("+", "").replace("-", "")
            # Add "-" before and after "Very Strong", "Strong", or "Weak"
            text = text.replace("Very Strong", "-Very Strong-").replace("Very Weak","-Very Weak-").replace("Strong", "-Strong-").replace("Weak", "-Weak-").replace("Very Weak","-Very Weak-")
            text = text.replace("Very -Strong-", "Very Strong").replace("Very -Weak-","Very Weak")
            # Split the text by "-" and write to the Excel file
            parts = text.split("-")
            for i in range(0, len(parts), 2):
                sheet.cell(row, 1, parts[i].strip())
                if i + 1 < len(parts):
                    sheet.cell(row, 2, parts[i + 1].strip())
                row += 1
    


    

# Save the Excel file
workbook.save("Strength and weakness.xlsx")

# Close the browser
driver.quit()

workbook = load_workbook("Strength and weakness.xlsx")

# Remove the "Sheet" sheet if it exists
if "Sheet" in workbook.sheetnames:
    workbook.remove(workbook["Sheet"])

# Rename the specific sheets
sheets_to_rename = ["Manchester-United", "Fulham", "Tottenham", "Manchester-City", "West-Ham",
                    "Newcastle", "Everton", "Nottingham-Forest", "Brentford", "Liverpool",
                    "Arsenal", "Sheffield-United", "Crystal-Palace", "Luton", "Burnley", "Brighton",
                    "Bournemouth", "Wolves", "Aston-Villa"]
new_sheet_names = ["Manchester United", "FC Fulham", "Tottenham Hotspur", "Manchester City", "West Ham United",
                   "Newcastle United", "FC Everton", "Nottingham Forest", "FC Brentford", "Liverpool", "Arsenal",
                   "Sheffield United", "Crystal Palace", "Luton Town", "FC Burnley", "Brighton & Hove Albion",
                   "AFC Bournemouth", "Wolverhampton Wanderers", "Aston Villa"]

for sheet_name, new_name in zip(sheets_to_rename, new_sheet_names):
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        new_rows = []
        for row in sheet.iter_rows(values_only=True):
            new_row = []
            for cell in row:
                if isinstance(cell, str):
                    # Remove words between brackets and brackets
                    cell = re.sub(r'\([^)]*\)', '', cell).strip()
                new_row.append(cell)
            new_rows.append(new_row)

        # Clear existing rows
        sheet.delete_rows(1, sheet.max_row)

        # Append modified rows
        for new_row in new_rows:
            sheet.append(new_row)

        sheet.title = new_name

# Save the changes to the Excel file
workbook.save("Strength and weakness.xlsx")


# In[3]:


import re
from openpyxl import load_workbook

# Load the Excel file
workbook = load_workbook("Strength and weakness.xlsx")

# Remove the "Sheet" sheet if it exists
if "Sheet" in workbook.sheetnames:
    workbook.remove(workbook["Sheet"])

# Rename the specific sheets
sheets_to_rename = ["Manchester-United", "Fulham", "Tottenham Hotspur", "Manchester-City", "West-Ham",
                    "Newcastle United", "Everton", "Nottingham-Forest", "FC Brentford", "Liverpool",
                    "Arsenal", "Sheffield-United", "Crystal-Palace", "Luton", "FC Burnley", "Brighton",
                    "Bournemouth", "Wolves", "Aston-Villa"]
new_sheet_names = ["Manchester United", "FC Fulham", "Tottenham Hotspur", "Manchester City", "West Ham United",
                   "Newcastle United", "FC Everton", "Nottingham Forest", "FC Brentford", "Liverpool", "Arsenal",
                   "Sheffield United", "Crystal Palace", "Luton Town", "FC Burnley", "Brighton & Hove Albion",
                   "AFC Bournemouth", "Wolverhampton Wanderers", "Aston Villa"]

for sheet_name, new_name in zip(sheets_to_rename, new_sheet_names):
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        new_rows = []
        for row in sheet.iter_rows(values_only=True):
            new_row = []
            for cell in row:
                if isinstance(cell, str):
                    # Remove words between brackets and brackets
                    cell = re.sub(r'\([^)]*\)', '', cell).strip()
                new_row.append(cell)
            new_rows.append(new_row)

        # Clear existing rows
        sheet.delete_rows(1, sheet.max_row)

        # Append modified rows
        for new_row in new_rows:
            sheet.append(new_row)

        sheet.title = new_name

# Save the changes to the Excel file
workbook.save("Strength and weakness.xlsx")


# In[1]:


import re
from openpyxl import load_workbook

workbook = load_workbook("Strength and weakness.xlsx")

# Remove the "Sheet" sheet if it exists
if "Sheet" in workbook.sheetnames:
    workbook.remove(workbook["Sheet"])

# Rename the specific sheets
sheets_to_rename = ["Manchester-United", "Fulham", "Tottenham", "Manchester-City", "West-Ham",
                    "Newcastle", "Everton", "Nottingham-Forest", "Brentford", "Liverpool",
                    "Arsenal", "Sheffield-United", "Crystal-Palace", "Luton", "Burnley", "Brighton",
                    "Bournemouth", "Wolves", "Aston-Villa"]
new_sheet_names = ["Manchester United", "FC Fulham", "Tottenham Hotspur", "Manchester City", "West Ham United",
                   "Newcastle United", "FC Everton", "Nottingham Forest", "FC Brentford", "Liverpool", "Arsenal",
                   "Sheffield United", "Crystal Palace", "Luton Town", "FC Burnley", "Brighton & Hove Albion",
                   "AFC Bournemouth", "Wolverhampton Wanderers", "Aston Villa"]

for sheet_name, new_name in zip(sheets_to_rename, new_sheet_names):
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        new_rows = []
        for row in sheet.iter_rows(values_only=True):
            new_row = []
            for cell in row:
                if isinstance(cell, str):
                    # Remove words between brackets and brackets
                    cell = re.sub(r'\([^)]*\)', '', cell).strip()
                new_row.append(cell)
            new_rows.append(new_row)

        # Clear existing rows
        sheet.delete_rows(1, sheet.max_row)

        # Append modified rows
        for new_row in new_rows:
            sheet.append(new_row)

        sheet.title = new_name

# Save the changes to the Excel file
workbook.save("Strength and weakness.xlsx")


# In[ ]:




