#!/usr/bin/env python
# coding: utf-8

# In[1]:


import csv
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver

# List of URLs containing the tables
urls = [
   
    'https://www.transfermarkt.com/fc-chelsea/startseite/verein/631',
 
]

chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'

# Replace with the path to your ChromeDriver executable
driver = webdriver.Chrome(executable_path=chrome_driver_path)

# Create an empty dictionary to store player data for each URL
all_player_data = {}

# Create a dictionary to map team names to specific sheet names
team_sheet_mapping = {

    'fc-chelsea': 'Chelsea',

}

for url in urls:
    # Load the page using Selenium
    driver.get(url)

    # Now, fetch the page content after a brief wait
    driver.implicitly_wait(5)
    page_content = driver.page_source

    # Parse the HTML content of the page
    soup = BeautifulSoup(page_content, 'html.parser')

    # Specify the table name
    table_name = 'items'

    # Find the table element based on its class name
    table = soup.find('table', class_='items')

    # Check if the table is found
    if table:
        # Create a list to store the data
        player_data = []

        # Extract and append player names based on positions
        positions = ["Goalkeeper", "Right Winger", "Left Winger", "Defensive Midfield", "Left-Back", "Attacking Midfield", "Centre-Back", "Right-Back", "Central Midfield", "Centre-Forward", "Second Striker", "Left Midfield", "Right Midfield"]

        for position in positions:
            for row in table.find_all('tr')[1:]:  # Skip the header row
                if position in row.get_text():
                    player_name_element = row.find('td', class_='hauptlink')
                    player_name = player_name_element.text.strip() if player_name_element else None
                    if player_name is not None:
                        player_data.append({"Player": player_name, "Position": position})

        # Store player data in the dictionary with URL as key
        all_player_data[url] = player_data
    else:
        print(f"No table with class '{table_name}' found on the page for URL: {url}")

# Close the browser window after processing all URLs
driver.quit()

# Save data to the same Excel file with different sheets
excel_file_path = 'playerseasondata2-23.xlsx'  # Use an Excel file for multiple sheets
with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
    for url, player_data in all_player_data.items():
        # Convert player data to a DataFrame
        df = pd.DataFrame(player_data, columns=["Player", "Position"])

        # Extract the team name from the URL
        team_name = url.split("/")[3]

        # Get the corresponding sheet name from the mapping dictionary
        sheet_name = team_sheet_mapping.get(team_name, team_name)

        # Write the DataFrame to a sheet with the specific name
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"Data saved to {excel_file_path}")

import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from unidecode import unidecode 
import openpyxl
import re

# URLs to scrape
urls = [

    'https://www.transfermarkt.com/fc-chelsea/startseite/verein/631',

    
    
]


chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'

# Replace with the path to your ChromeDriver executable
driver = webdriver.Chrome(executable_path=chrome_driver_path)

# Initialize an empty dictionary to store fifth column data for each sheet
fifth_column_data_dict = {}
sixth_column_data_dict = {}
seventh_column_data_dict = {}
eigth_column_data_dict = {}
nineth_column_data_dict = {}
tenth_column_data_dict = {}
eleventh_column_data_dict = {}
twelveth_column_data_dict = {}
thriteenth_column_data_dict = {}

# Iterate through each URL
for index, url in enumerate(urls):
    # Open the URL
    driver.get(url)

    # Get the HTML content of the page
    html = driver.page_source

    # Use BeautifulSoup to parse the HTML
    soup = BeautifulSoup(html, 'html.parser')

    # Find the table with class name "items"
    items_table = soup.find('table', class_='items')

    # Initialize an empty list to store filtered href links
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue

    # Read names from Excel workbook
    excel_path = 'playerseasondata2-23.xlsx'  # Replace with the actual path to your Excel file
    try:
        # Dynamically get the sheet name based on the index
        sheet_name = pd.ExcelFile(excel_path).sheet_names[index]
        df = pd.read_excel(excel_path, sheet_name=sheet_name, usecols=[0])  # Assuming names are in the second column
    except IndexError:
        print(f"No sheet found in the Excel file for URL {url}.")
        continue

    # Initialize an empty list to store fifth column data for the current sheet
    fifth_column_data_list = []
    sixth_column_data_list = []
    seventh_column_data_list = []
    eigth_column_data_list = []
    nineth_column_data_list = []
    tenth_column_data_list = []
    eleventh_column_data_list = []
    twelveth_column_data_list = []
    thriteenth_column_data_list = []

    # Iterate through names in the DataFrame
    for given_name in df['Player']:  # Assuming the column name is 'Column2'

        normalized_name = unidecode(given_name)
        
        normalized_name = normalized_name.replace("-"," ")
        
        normalized_name = re.sub(r'[^\w\s]', '', normalized_name)
        # Split the name and convert to lowercase
        name_parts = [part.lower() for part in normalized_name.split()]

        # Check if all parts are present in any element of filtered_href_list
        matching_elements = [element for element in filtered_href_list if all(part in element.lower() for part in name_parts)]

        # Save the matching elements in a variable
        saved_matching_elements = []

        if matching_elements:
            for element in matching_elements:
                saved_matching_elements.append(element)
        else:
            print(f"No matching elements found for {given_name} in sheet {sheet_name}.")

        if saved_matching_elements:
            split_elements = [element.split('/') for element in saved_matching_elements]

            first_list = split_elements[0]

            new_variable = 'https://www.transfermarkt.com/' + first_list[1] + '/leistungsdaten/spieler/' + first_list[-1]+ "/plus/0?saison=2022"

            driver.get(new_variable)
            # Get the HTML content of the new page
            html_new_page = driver.page_source

            # Use BeautifulSoup to parse the HTML of the new page
            soup_new_page = BeautifulSoup(html_new_page, 'html.parser')

            # Find the table containing the performance data
            table = soup_new_page.find('table', {'class': 'items'})
            
            ele = soup_new_page.find_all(class_="data-header__content")

                    # Extract the text from the found elements
            tex1 = [element.get_text() for element in ele]

                    # Check if any text contains "Goalkeeper"
            contains_goalkeeper = any("Goalkeeper" in text for text in tex1)

            # Check if the table is found
            if table:
                # Find the tfoot element
                tfoot = table.find('tfoot')

                # Check if tfoot is found
                if tfoot:
                    # Extract data from the tfoot
                    rows = tfoot.find_all('tr')

                    for row in rows:
                        # Extract data from the fifth column (index 4) of each row
                        columns = row.find_all(['td', 'th'])
                        fifth_column_data = columns[2].get_text(strip=True)

                        # Append the fifth column data to the list
                        fifth_column_data_list.append(fifth_column_data)
                    
                    for row in rows:
                        # Extract data from the fifth column (index 5) of each row
                        columns = row.find_all(['td', 'th'])
                        sixth_column_data = columns[3].get_text(strip=True)

                        # Append the fifth column data to the list
                        sixth_column_data_list.append(sixth_column_data)
                    

                    
                    if contains_goalkeeper:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            eigth_column_data = columns[7].get_text(strip=True)

                            # Append the fifth column data to the list
                            eigth_column_data_list.append(eigth_column_data)
                      
                    else:
                        eigth_column_data_list.append("-")
                        
                    if contains_goalkeeper:
                        seventh_column_data_list.append("-")
                                            
                    else:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            
                            # Extract data from the fifth column (index 5) of each row
                            seventh_column_data = columns[4].get_text(strip=True)
                            seventh_column_data_list.append(seventh_column_data)
                            
                    if contains_goalkeeper:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            nineth_column_data = columns[8].get_text(strip=True)

                            # Append the fifth column data to the list
                            nineth_column_data_list.append(nineth_column_data)
                      
                    else:
                        nineth_column_data_list.append("-")
                        
                    if contains_goalkeeper:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            tenth_column_data = columns[4].get_text(strip=True)

                            # Append the fifth column data to the list
                            tenth_column_data_list.append(tenth_column_data)
                      
                    else:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            tenth_column_data = columns[5].get_text(strip=True)

                            # Append the fifth column data to the list
                            tenth_column_data_list.append(tenth_column_data)
                            
                    if contains_goalkeeper:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            eleventh_column_data = columns[6].get_text(strip=True)

                            # Append the fifth column data to the list
                            eleventh_column_data_list.append(eleventh_column_data)
                      
                    else:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            eleventh_column_data = columns[7].get_text(strip=True)

                            # Append the fifth column data to the list
                            eleventh_column_data_list.append(eleventh_column_data)
                            
                    if contains_goalkeeper:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            twelveth_column_data = columns[5].get_text(strip=True)

                            # Append the fifth column data to the list
                            twelveth_column_data_list.append(eleventh_column_data)
                      
                    else:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            twelveth_column_data = columns[6].get_text(strip=True)

                            # Append the fifth column data to the list
                            twelveth_column_data_list.append(twelveth_column_data)
                            
                    if contains_goalkeeper:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            thriteenth_column_data = columns[9].get_text(strip=True)

                            # Append the fifth column data to the list
                            thriteenth_column_data_list.append(thriteenth_column_data)
                      
                    else:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            thriteenth_column_data = columns[8].get_text(strip=True)

                            # Append the fifth column data to the list
                            thriteenth_column_data_list.append(thriteenth_column_data)
                    
                        
                else:
                    print(f"No tfoot found in the table for {given_name} in sheet {sheet_name}.")
            else:
                fifth_column_data_list.append("-")
                sixth_column_data_list.append("-")
                seventh_column_data_list.append("-")
                eigth_column_data_list.append("-")
                nineth_column_data_list.append("-")
                tenth_column_data_list.append("-")
                eleventh_column_data_list.append("-")
                twelveth_column_data_list.append("-")
                thriteenth_column_data_list.append("-")
                print(f"No table found on the page for {given_name} in sheet {sheet_name}.")

    # Store fifth column data list for the current sheet in the dictionary
    fifth_column_data_dict[sheet_name] = fifth_column_data_list
    sixth_column_data_dict[sheet_name] = sixth_column_data_list
    seventh_column_data_dict[sheet_name] = seventh_column_data_list
    eigth_column_data_dict[sheet_name] = eigth_column_data_list
    nineth_column_data_dict[sheet_name] = nineth_column_data_list
    tenth_column_data_dict[sheet_name] = tenth_column_data_list
    eleventh_column_data_dict[sheet_name] = eleventh_column_data_list
    twelveth_column_data_dict[sheet_name] = twelveth_column_data_list
    thriteenth_column_data_dict[sheet_name] = thriteenth_column_data_list

# Close the webdriver
driver.quit()

# Load the workbook
workbook = openpyxl.load_workbook('playerseasondata2-23.xlsx')

# Access the sheets in the workbook
sheets = workbook.sheetnames

# Your dictionary

# Loop through each sheet and add the data
for sheet_name, values in zip(sheets, fifth_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['c1'] = "Season appearance"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=3, value=value)
        
# Loop through each sheet and add the data
for sheet_name, values in zip(sheets, sixth_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['d1'] = "Goals"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=4, value=value)
        
for sheet_name, values in zip(sheets, seventh_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['e1'] = "Assists"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=5, value=value)
        
for sheet_name, values in zip(sheets, eigth_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['f1'] = "Goal conceded"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=6, value=value)
        
for sheet_name, values in zip(sheets, nineth_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['g1'] = "Clean sheet"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=7, value=value)
        
for sheet_name, values in zip(sheets, tenth_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['h1'] = "Yellow cards"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=8, value=value)
        
for sheet_name, values in zip(sheets, eleventh_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['i1'] = "Red cards"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=9, value=value)
        
for sheet_name, values in zip(sheets, twelveth_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['j1'] = "Second yellow cards"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=10, value=value)
        
for sheet_name, values in zip(sheets, thriteenth_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['k1'] = "Played minutes"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=11, value=value)



# Save the workbook
workbook.save('playerseasondata2-23.xlsx')

import openpyxl
workbook = openpyxl.load_workbook('playerseasondata2-23.xlsx')

# Access the sheets in the workbook
sheets = workbook.sheetnames

# Loop through each sheet and update the "Played minutes" column
for sheet_name in sheets:
    sheet = workbook[sheet_name]
    # Iterate over each cell in the "Played minutes" column starting from the second row
    for row in sheet.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            # Get the current value of the cell
            current_value = cell.value
            # Remove "." and "'" from the current value
            if current_value is not None:
                updated_value = current_value.replace(".", "").replace("'", "")
                # Update the cell value
                cell.value = updated_value

# Save the workbook
workbook.save('playerseasondata2-23.xlsx')


# In[2]:





# In[3]:





# In[ ]:




