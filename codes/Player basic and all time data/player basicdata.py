import csv
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver


urls = [

    'https://www.transfermarkt.com/fc-chelsea/startseite/verein/631']


chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'

#  ChromeDriver 
driver = webdriver.Chrome(executable_path=chrome_driver_path)

# Create an empty dictionary to store player data for each URL
all_player_data = {}

# Create a dictionary to map team names to specific sheet names
team_sheet_mapping = {

    'fc-chelsea': 'Chelsea',

}

for url in urls:
    # Load the page using Selenium
    driver.get(url)

    # Now, fetch the page content after a brief wait
    driver.implicitly_wait(5)
    page_content = driver.page_source

    # Parse the HTML content of the page
    soup = BeautifulSoup(page_content, 'html.parser')

    # Specify the table name
    table_name = 'items'

    # Find the table element based on its class name
    table = soup.find('table', class_='items')

    # Check if the table is found
    if table:
        # Create a list to store the data
        player_data = []

        # Extract and append player names based on positions
        positions = ["Goalkeeper", "Right Winger", "Left Winger", "Defensive Midfield", "Left-Back", "Attacking Midfield", "Centre-Back", "Right-Back", "Central Midfield", "Centre-Forward", "Second Striker", "Left Midfield", "Right Midfield"]

        for position in positions:
            for row in table.find_all('tr')[1:]:  # Skip the header row
                if position in row.get_text():
                    player_name_element = row.find('td', class_='hauptlink')
                    player_name = player_name_element.text.strip() if player_name_element else None
                    if player_name is not None:
                        player_data.append({"Player": player_name, "Position": position})

        # Store player data in the dictionary with URL as key
        all_player_data[url] = player_data
    else:
        print(f"No table with class '{table_name}' found on the page for URL: {url}")

# Close the browser window 
driver.quit()

# Save data to the same Excel file with different sheets
excel_file_path = 'player_data.xlsx'  
with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
    for url, player_data in all_player_data.items():
        # Convert player data to a DataFrame
        df = pd.DataFrame(player_data, columns=["Player", "Position"])

        # Extract the team name from the URL
        team_name = url.split("/")[3]

        # Get the corresponding sheet name from the mapping dictionary
        sheet_name = team_sheet_mapping.get(team_name, team_name)

        
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"Data saved to {excel_file_path}")
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from unidecode import unidecode 
import openpyxl
import re
import time
# URLs to scrape
urls = [ 
    'https://www.transfermarkt.com/fc-chelsea/startseite/verein/631'
       ]

chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'


driver = webdriver.Chrome(executable_path=chrome_driver_path)

# Initialize an empty dictionary to store final values for each sheet
final_values_dict = {}
final_hiegth_dict = {}
final_num_dict = {}
final_age_dict = {}
final_foot_dict = {}

# Iterate through each URL
for index, url in enumerate(urls):
    # Open the URL
    driver.get(url)

    # Get the HTML content of the page
    html = driver.page_source

    # Use BeautifulSoup to parse the HTML
    soup = BeautifulSoup(html, 'html.parser')

    # Find the table with class name "items"
    items_table = soup.find('table', class_='items')

    # Initialize an empty list to store filtered href links
    filtered_href_list = []

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue
        

    # Read names from Excel workbook
    excel_path = 'player_data.xlsx'  # Replace with the actual path to your Excel file
    try:
        # Dynamically get the sheet name based on the index
        sheet_name = pd.ExcelFile(excel_path).sheet_names[index]
        df = pd.read_excel(excel_path, sheet_name=sheet_name, usecols=[0])  # Assuming names are in the first column
    except IndexError:
        print(f"No sheet found in the Excel file for URL {url}.")
        continue

    # Initialize a list to store final values for this sheet
    final_values_list = []
    final_hiegth_list =[]
    final_num_list =[]
    final_age_list =[]
    final_foot_list =[]

    # Iterate through names in the DataFrame
    for given_name in df['Player']:  # Assuming the column name is 'Player'
        normalized_name = unidecode(given_name)
        
        normalized_name = re.sub(r'[^\w\s]', '', normalized_name)
        # Split the name and convert to lowercase
        name_parts = [part.lower() for part in normalized_name.split()]

        # Check if all parts are present in any element of filtered_href_list
        matching_elements = [element for element in filtered_href_list if all(part in element.lower() for part in name_parts)]

        # Prepend "https://www.transfermarkt.com/" to each matching element
        matching_elements = ['https://www.transfermarkt.com' + element for element in matching_elements]
        

        # Iterate through matching elements for each player
        for i in matching_elements:
            
            driver.get(i)
            
            html = driver.page_source
            
            soup = BeautifulSoup(html, 'html.parser')
            time.sleep(5)
            # Find market value element
            elements = soup.find_all(class_='detail-position')
            info_table = soup.find('div', class_='info-table')
            
            for element in elements:
                text = element.get_text()
                if "Other position:" in text:
                    other_position_index = text.index("Other position:")
                    other_position_text = text[other_position_index + len("Other position:"):].strip()
                    final_values_list.append(other_position_text)
                else:
                    final_values_list.append("-")
                    
            try:
                height_element = soup.find('span', itemprop='height')
                height_text = height_element.get_text()
                height_text = height_text.replace(' ', '').replace(',', '').replace('m', 'cm')
                final_hiegth_list.append(height_text)
                
            except:
                final_hiegth_list.append("-")
                
            try:
                num_element = soup.find('span',class_='data-header__shirt-number')
                num_text = num_element.get_text()
                num_text = num_text.replace('#', '')
                final_num_list.append(num_text)
                
            except:
                final_num_list.append("-")
                
            try:
                age_element = soup.find('span', itemprop='birthDate')
                age_text = age_element.get_text()
                final_age_list.append(age_text)
                
            except:
                final_age_list.append("-")
                
            try:
                if info_table:
                    spans = info_table.find_all('span', class_='info-table__content')
                     
                    if spans[12].get_text(strip=True)== "Foot:":
                        duminant_foot_row = spans[13].get_text(strip=True)
                        final_foot_list.append(duminant_foot_row)
                    else:
                        duminant_foot_row = spans[11].get_text(strip=True)
                        final_foot_list.append(duminant_foot_row)
                
            except:
                final_foot_list.append("-")
            

    # Add the list of final values for this sheet to the dictionary
    final_values_dict[sheet_name] = final_values_list
    final_hiegth_dict[sheet_name] = final_hiegth_list
    final_num_dict[sheet_name] = final_num_list
    final_age_dict[sheet_name] = final_age_list
    final_foot_dict[sheet_name] = final_foot_list
    

# Close the webdriver
driver.quit()

workbook = openpyxl.load_workbook('player_data.xlsx')

# Access the sheets in the workbook
sheets = workbook.sheetnames


# Loop through each sheet and add the data
for sheet_name, values in zip(sheets, final_values_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['c1'] = "Other Positions"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=3, value=value)
        
for sheet_name, values in zip(sheets, final_hiegth_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['d1'] = "Height"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=4, value=value)
        
for sheet_name, values in zip(sheets, final_num_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['e1'] = "Jersey num"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=5, value=value)
        
for sheet_name, values in zip(sheets, final_age_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['g1'] = "Date of birth/Age"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=7, value=value)
        
        
for sheet_name, values in zip(sheets, final_foot_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['h1'] = "Dominant foot"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=8, value=value)
        

        
workbook.save('player_data.xlsx')

import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from unidecode import unidecode 
import openpyxl
import re
# URLs to scrape
urls = [ 
    'https://www.transfermarkt.com/fc-chelsea/startseite/verein/631'
       ]

chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'


driver = webdriver.Chrome(executable_path=chrome_driver_path)

# Initialize an empty dictionary to store final values for each sheet
final_values_dict = {}


for index, url in enumerate(urls):
    # Open the URL
    driver.get(url)

    # Get the HTML content of the page
    html = driver.page_source

    # Use BeautifulSoup to parse the HTML
    soup = BeautifulSoup(html, 'html.parser')

    # Find the table with class name "items"
    items_table = soup.find('table', class_='items')

    # Initialize an empty list to store filtered href links
    filtered_href_list = []

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue

    # Read names from Excel workbook
    excel_path = 'player_data.xlsx'  # Replace with the actual path to your Excel file
    try:
        # Dynamically get the sheet name based on the index
        sheet_name = pd.ExcelFile(excel_path).sheet_names[index]
        df = pd.read_excel(excel_path, sheet_name=sheet_name, usecols=[0])  # Assuming names are in the first column
    except IndexError:
        print(f"No sheet found in the Excel file for URL {url}.")
        continue

    # Initialize a list to store final values for this sheet
    final_values_list = []

    # Iterate through names in the DataFrame
    for given_name in df['Player']:  # Assuming the column name is 'Player'
        normalized_name = unidecode(given_name)
        
        normalized_name = re.sub(r'[^\w\s]', '', normalized_name)
        # Split the name and convert to lowercase
        name_parts = [part.lower() for part in normalized_name.split()]

        # Check if all parts are present in any element of filtered_href_list
        matching_elements = [element for element in filtered_href_list if all(part in element.lower() for part in name_parts)]

        # Prepend "https://www.transfermarkt.com/" to each matching element
        matching_elements = ['https://www.transfermarkt.com' + element for element in matching_elements]
        
        # Iterate through matching elements for each player
        for element in matching_elements:
            driver.get(element)
            
            html = driver.page_source
            
            soup = BeautifulSoup(html, 'html.parser')
            
            # Find market value element
            market_value_element = soup.find(class_="data-header__market-value-wrapper")
            
            if market_value_element:
                # Extract value and denomination
                span_element = market_value_element.find('span')
                if span_element:
                    value = span_element.next_sibling.strip()
                    denomination = span_element.next_sibling.next_sibling.text
                    # Concatenate the value and the denomination
                    final_value = f"{value}{denomination}"
                    # Print the final value
                    
                    # Append the final value to the list for this sheet
                    final_values_list.append(final_value)
                else:
                    
                    print("Market value not found for", given_name)
            else:
                final_values_list.append("-")
                print("Market value not found for", given_name)

    # Add the list of final values for this sheet to the dictionary
    final_values_dict[sheet_name] = final_values_list
    

# Close the webdriver
driver.quit()

workbook = openpyxl.load_workbook('player_data.xlsx')

# Access the sheets in the workbook
sheets = workbook.sheetnames

# Your dictionary

# Loop through each sheet and add the data
for sheet_name, values in zip(sheets, final_values_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['f1'] = "Market value"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=6, value=value)
        
workbook.save('player_data.xlsx')


# In[10]:





# In[12]:





# In[ ]:




