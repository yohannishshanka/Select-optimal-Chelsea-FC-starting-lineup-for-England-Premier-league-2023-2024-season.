import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from unidecode import unidecode 
import openpyxl
import re

# URLs to scrape
urls = [

    'https://www.transfermarkt.com/fc-chelsea/startseite/verein/631'
    
]


chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'

driver = webdriver.Chrome(executable_path=chrome_driver_path)

# Initialize an empty dictionary to store fifth column data for each sheet
fifth_column_data_dict = {}
sixth_column_data_dict = {}
seventh_column_data_dict = {}
eigth_column_data_dict = {}
nineth_column_data_dict = {}

# Iterate through each URL
for index, url in enumerate(urls):
    # Open the URL
    driver.get(url)

    # Get the HTML content of the page
    html = driver.page_source

    # Use BeautifulSoup to parse the HTML
    soup = BeautifulSoup(html, 'html.parser')

    # Find the table with class name "items"
    items_table = soup.find('table', class_='items')

    # Initialize an empty list to store filtered href links
    filtered_href_list = []
    

    # Find all href links within the "items" table
    if items_table:
        href_links = items_table.find_all('a', href=True)

        # Extract and append only the href links containing "profil/spieler" to the list
        for link in href_links:
            href = link['href']
            if 'profil/spieler' in href:
                filtered_href_list.append(href)
    else:
        print(f"Table with class 'items' not found on the page for URL {url}.")
        continue

    # Read names from Excel workbook
    excel_path = 'player_data.xlsx'  
    try:
        # Dynamically get the sheet name based on the index
        sheet_name = pd.ExcelFile(excel_path).sheet_names[index]
        df = pd.read_excel(excel_path, sheet_name=sheet_name, usecols=[0])  # Assuming names are in the second column
    except IndexError:
        print(f"No sheet found in the Excel file for URL {url}.")
        continue

    # Initialize an empty list to store fifth column data for the current sheet
    fifth_column_data_list = []
    sixth_column_data_list = []
    seventh_column_data_list = []
    eigth_column_data_list = []
    nineth_column_data_list = []

    # Iterate through names in the DataFrame
    for given_name in df['Player']:  # Assuming the column name is 'Column2'

        normalized_name = unidecode(given_name)
        
        normalized_name = re.sub(r'[^\w\s]', '', normalized_name)
        # Split the name and convert to lowercase
        name_parts = [part.lower() for part in normalized_name.split()]

        # Check if all parts are present in any element of filtered_href_list
        matching_elements = [element for element in filtered_href_list if all(part in element.lower() for part in name_parts)]

        # Save the matching elements in a variable
        saved_matching_elements = []

        if matching_elements:
            for element in matching_elements:
                saved_matching_elements.append(element)
        else:
            print(f"No matching elements found for {given_name} in sheet {sheet_name}.")

        if saved_matching_elements:
            split_elements = [element.split('/') for element in saved_matching_elements]

            first_list = split_elements[0]

            new_variable = 'https://www.transfermarkt.com/' + first_list[1] + '/detaillierteleistungsdaten/spieler/' + first_list[-1]

            driver.get(new_variable)
            # Get the HTML content of the new page
            html_new_page = driver.page_source

            # Use BeautifulSoup to parse the HTML of the new page
            soup_new_page = BeautifulSoup(html_new_page, 'html.parser')

            # Find the table containing the performance data
            table = soup_new_page.find('table', {'class': 'items'})
            
            ele = soup_new_page.find_all(class_="data-header__content")

                    # Extract the text from the found elements
            tex1 = [element.get_text() for element in ele]

                    # Check if any text contains "Goalkeeper"
            contains_goalkeeper = any("Goalkeeper" in text for text in tex1)

            # Check if the table is found
            if table:
                # Find the tfoot element
                tfoot = table.find('tfoot')

             
                if tfoot:
                   
                    rows = tfoot.find_all('tr')

                    for row in rows:
                        
                        columns = row.find_all(['td', 'th'])
                        fifth_column_data = columns[4].get_text(strip=True)

                        
                        fifth_column_data_list.append(fifth_column_data)
                    
                    for row in rows:
                        
                        columns = row.find_all(['td', 'th'])
                        sixth_column_data = columns[5].get_text(strip=True)

                        
                        sixth_column_data_list.append(sixth_column_data)
                    

                    
                    if contains_goalkeeper:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            eigth_column_data = columns[6].get_text(strip=True)

                            
                            eigth_column_data_list.append(eigth_column_data)
                      
                    else:
                        eigth_column_data_list.append("-")
                        
                    if contains_goalkeeper:
                        seventh_column_data_list.append("-")
                                            
                    else:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            
                            
                            seventh_column_data = columns[6].get_text(strip=True)
                            seventh_column_data_list.append(seventh_column_data)
                            
                    if contains_goalkeeper:
                        for row in rows:
                            columns = row.find_all(['td', 'th'])
                            nineth_column_data = columns[7].get_text(strip=True)

                            
                            nineth_column_data_list.append(nineth_column_data)
                      
                    else:
                        nineth_column_data_list.append("-")
                    
                        
                else:
                    print(f"No tfoot found in the table for {given_name} in sheet {sheet_name}.")
            else:
                fifth_column_data_list.append("-")
                sixth_column_data_list.append("-")
                seventh_column_data_list.append("-")
                eigth_column_data_list.append("-")
                nineth_column_data_list.append("-")
                print(f"No table found on the page for {given_name} in sheet {sheet_name}.")

    
    fifth_column_data_dict[sheet_name] = fifth_column_data_list
    sixth_column_data_dict[sheet_name] = sixth_column_data_list
    seventh_column_data_dict[sheet_name] = seventh_column_data_list
    eigth_column_data_dict[sheet_name] = eigth_column_data_list
    nineth_column_data_dict[sheet_name] = nineth_column_data_list

# Close the webdriver
driver.quit()

# Load the workbook
workbook = openpyxl.load_workbook('player_data.xlsx')

# Access the sheets in the workbook
sheets = workbook.sheetnames

# Your dictionary

# Loop through each sheet and add the data
for sheet_name, values in zip(sheets, fifth_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['i1'] = "Appearance"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=9, value=value)
        
# Loop through each sheet and add the data
for sheet_name, values in zip(sheets, sixth_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['j1'] = "Goals"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=10, value=value)
        
for sheet_name, values in zip(sheets, seventh_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['k1'] = "Assists"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=11, value=value)
        
for sheet_name, values in zip(sheets, eigth_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['l1'] = "Goal conceded"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=12, value=value)
        
for sheet_name, values in zip(sheets, nineth_column_data_dict.values()):
    sheet = workbook[sheet_name]
    # Add column header
    sheet['m1'] = "Clean sheet"
    # Add values as a new column
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=13, value=value)


# Save the workbook
workbook.save('player_data.xlsx')

from selenium import webdriver
from bs4 import BeautifulSoup
import openpyxl

def scrape_and_update(url):
    # Using Selenium to fetch the page source
    chrome_driver_path = 'C:/Users/Admin/Documents/research final - chelsea/Research/chromedriver.exe'


    driver = webdriver.Chrome(executable_path=chrome_driver_path)

    driver.get(url)

    # Extracting page source
    page_source = driver.page_source

    # Close the Selenium driver
    driver.quit()

    # Parsing the page source with BeautifulSoup
    soup = BeautifulSoup(page_source, "html.parser")

    # Find all <a> elements within the table data
    a_elements = soup.find_all("td", class_="hauptlink")  # Adjust the class name as needed

    # Dictionary to store player names as keys and injury information as values
    player_injury_dict = {}

    # Extract player names and injury information (if present)
    for a in a_elements:
        player_name = a.text.strip()  # Extract player name
        span_title = a.find("span", title=True)  # Find <span> element with title attribute
        if span_title:
            injury_info = span_title["title"]
            if injury_info != "Team captain":
                player_injury_dict[player_name] = injury_info

    # Load workbook
    workbook = openpyxl.load_workbook('player_data.xlsx')
    
    sheet = workbook['Chelsea']
    
    sheet.cell(row=1, column=14, value='Injury update')

    # Iterate over each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Iterate over each row in the sheet
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                player_name = cell.value
                if player_name in player_injury_dict:
                    # If the player name is found in the dictionary, insert the injury information into the 9th column
                    sheet.cell(row=cell.row, column=14, value=player_injury_dict[player_name])

    # Save the workbook
    workbook.save('player_data.xlsx')

# List of URLs to scrape
urls = [

    'https://www.transfermarkt.com/fc-chelsea/startseite/verein/631',
 
]

# Loop through each URL and apply the scraping and updating function
for url in urls:
    scrape_and_update(url)
